import os
import json
import hashlib
from pathlib import Path

import numpy as np
import faiss
import ollama
import tiktoken

from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook
import csv


CHUNK_SIZE = 500
CHUNK_OVERLAP = 50
EMBED_MODEL = "nomic-embed-text"


def extract_text_from_pdf(filepath: str) -> str:
    reader = PdfReader(filepath)
    texts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            texts.append(text)
    return "\n".join(texts)


def extract_text_from_docx(filepath: str) -> str:
    doc = DocxDocument(filepath)
    return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])


def extract_text_from_xlsx(filepath: str) -> str:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    texts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        texts.append(f"--- Sheet: {sheet} ---")
        for row in ws.iter_rows(values_only=True):
            row_str = " | ".join(str(cell) for cell in row if cell is not None)
            if row_str.strip():
                texts.append(row_str)
    wb.close()
    return "\n".join(texts)


def extract_text_from_csv(filepath: str) -> str:
    texts = []
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                texts.append("Headers: " + " | ".join(row))
            else:
                texts.append(" | ".join(row))
    return "\n".join(texts)


def extract_text_from_txt(filepath: str) -> str:
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def load_document(filepath: str) -> str:
    ext = Path(filepath).suffix.lower()
    loaders = {
        ".pdf": extract_text_from_pdf,
        ".docx": extract_text_from_docx,
        ".xlsx": extract_text_from_xlsx,
        ".csv": extract_text_from_csv,
        ".txt": extract_text_from_txt,
    }
    loader = loaders.get(ext)
    if not loader:
        raise ValueError(f"Unsupported file format: {ext}")
    return loader(filepath)


def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    enc = tiktoken.get_encoding("cl100k_base")
    tokens = enc.encode(text)
    chunks = []
    stride = chunk_size - overlap
    for i in range(0, len(tokens), stride):
        chunk_tokens = tokens[i : i + chunk_size]
        chunk = enc.decode(chunk_tokens)
        if chunk.strip():
            chunks.append(chunk)
    return chunks


def embed_texts(texts: list[str], model: str = EMBED_MODEL) -> list[list[float]]:
    embeddings = []
    for text in texts:
        resp = ollama.embed(model=model, input=text)
        embeddings.extend(resp["embeddings"])
    return embeddings


def file_hash(filepath: str) -> str:
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def load_index(store_dir: str):
    index_path = os.path.join(store_dir, "index.faiss")
    docs_path = os.path.join(store_dir, "docs.json")

    if os.path.exists(index_path) and os.path.exists(docs_path):
        index = faiss.read_index(index_path)
        with open(docs_path, "r") as f:
            docs = json.load(f)
        return index, docs
    return None, []


def save_index(store_dir: str, index, docs: list[dict]):
    os.makedirs(store_dir, exist_ok=True)
    faiss.write_index(index, os.path.join(store_dir, "index.faiss"))
    with open(os.path.join(store_dir, "docs.json"), "w") as f:
        json.dump(docs, f)


def build_index_from_docs(docs: list[dict]) -> faiss.Index:
    if not docs:
        dim = 768
        index = faiss.IndexFlatIP(dim)
        return index

    embeddings = np.array([d["embedding"] for d in docs], dtype=np.float32)
    dim = embeddings.shape[1]

    norms = np.linalg.norm(embeddings, axis=1, keepdims=True)
    norms[norms == 0] = 1e-9
    embeddings = embeddings / norms

    index = faiss.IndexFlatIP(dim)
    index.add(embeddings)
    return index


def ingest_files(filepaths: list[str], store_dir: str, replace: bool = False) -> dict:
    os.makedirs(store_dir, exist_ok=True)

    if replace:
        index_path = os.path.join(store_dir, "index.faiss")
        docs_path = os.path.join(store_dir, "docs.json")
        for p in [index_path, docs_path]:
            if os.path.exists(p):
                os.remove(p)

    index, docs = load_index(store_dir)
    existing_hashes = {d.get("file_hash", "") for d in docs}

    results = {"processed": 0, "skipped": 0, "errors": 0, "chunks": 0}
    new_docs = []

    for filepath in filepaths:
        if not os.path.exists(filepath):
            results["errors"] += 1
            continue

        fhash = file_hash(filepath)
        if fhash in existing_hashes and not replace:
            results["skipped"] += 1
            continue

        try:
            text = load_document(filepath)
            if not text.strip():
                results["skipped"] += 1
                continue

            chunks = chunk_text(text)
            if not chunks:
                results["skipped"] += 1
                continue

            embeddings = embed_texts(chunks)

            for i, (chunk, emb) in enumerate(zip(chunks, embeddings)):
                new_docs.append({
                    "id": f"{fhash}_{i}",
                    "text": chunk,
                    "embedding": emb,
                    "filename": os.path.basename(filepath),
                    "filepath": filepath,
                    "file_hash": fhash,
                    "chunk_index": i,
                })

            results["processed"] += 1
            results["chunks"] += len(chunks)
        except Exception as e:
            print(f"Error processing {filepath}: {e}")
            results["errors"] += 1

    if new_docs:
        all_docs = docs + new_docs
        index = build_index_from_docs(all_docs)
        save_index(store_dir, index, all_docs)

    return results


def ingest_folder(folder_path: str, store_dir: str, replace: bool = False) -> dict:
    supported_ext = {".pdf", ".docx", ".xlsx", ".csv", ".txt"}
    filepaths = []
    for root, _, files in os.walk(folder_path):
        for fname in files:
            if Path(fname).suffix.lower() in supported_ext:
                filepaths.append(os.path.join(root, fname))

    return ingest_files(filepaths, store_dir, replace)
